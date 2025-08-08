from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle

class ExcelStyler:
    """Clase para manejar estilos del Excel de forma profesional"""

    def __init__(self):
        # Definir paleta de colores corporativa
        self.colors = {
            'primary': 'FF2B5CE6',      # Azul principal
            'secondary': 'FF10B981',    # Verde éxito
            'accent': 'FFFBBF24',       # Amarillo acento
            'danger': 'FFEF4444',       # Rojo error
            'dark': 'FF1F2937',         # Gris oscuro
            'light': 'FFF8FAFC',        # Gris claro
            'white': 'FFFFFFFF',        # Blanco
            'header_bg': 'FF1E40AF',    # Azul header
            'alt_row': 'FFF1F5F9',      # Fila alternativa
            'neto_positive': 'FF10B981', # Verde para neto positivo
            'neto_negative': 'FFEF4444', # Rojo para neto negativo
            'haberes_bg': 'FF10B981',   # Verde para sección haberes
            'deducciones_bg': 'FFEF4444', # Rojo para sección deducciones
            'totales_bg': 'FF6366F1'    # Púrpura para totales
        }

    def create_header_style(self):
        """Estilo para headers de tabla"""
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(
            name='Calibri',
            size=12,
            bold=True,
            color='FFFFFFFF'
        )
        header_style.fill = PatternFill(
            start_color=self.colors['header_bg'],
            end_color=self.colors['header_bg'],
            fill_type='solid'
        )
        header_style.border = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thick', color='FF000000')
        )
        header_style.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        return header_style

    def create_data_style(self):
        """Estilo para datos de tabla"""
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(
            name='Calibri',
            size=11,
            color='FF1F2937'
        )
        data_style.border = Border(
            left=Side(style='thin', color='FFE5E7EB'),
            right=Side(style='thin', color='FFE5E7EB'),
            top=Side(style='thin', color='FFE5E7EB'),
            bottom=Side(style='thin', color='FFE5E7EB')
        )
        data_style.alignment = Alignment(
            horizontal='right',
            vertical='center'
        )
        data_style.number_format = '$#,##0.00'
        return data_style

    def create_period_style(self):
        """Estilo para columna de períodos"""
        period_style = NamedStyle(name="period_style")
        period_style.font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color='FF1E40AF'
        )
        period_style.fill = PatternFill(
            start_color=self.colors['light'],
            end_color=self.colors['light'],
            fill_type='solid'
        )
        period_style.border = Border(
            left=Side(style='thick', color='FF1E40AF'),
            right=Side(style='thin', color='FFE5E7EB'),
            top=Side(style='thin', color='FFE5E7EB'),
            bottom=Side(style='thin', color='FFE5E7EB')
        )
        period_style.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        return period_style
