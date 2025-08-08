import tkinter as tk
from tkinter import messagebox, filedialog
import customtkinter as ctk
import threading
import time
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
from PyPDF2 import PdfMerger
import glob
import subprocess
import platform
from datetime import datetime
import hashlib
import getpass
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from excel_styler import ExcelStyler
from license_manager import LicenseManager, LicenseDialog
from console_widget import ConsoleWidget

# Configurar tema de customtkinter
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class ANSESDownloaderPro:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.iconbitmap("icono_recibos.ico")
        self.root.title("Descargador de Recibos de Anses")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)
        
        # Variables
        self.driver = None
        self.is_running = False
        self.pdf_final_path = ""
        self.animation_running = False
        self.license_manager = LicenseManager()
        
        # NUEVA VARIABLE: Almacenar TODOS los datos de períodos
        self.todos_los_datos = []  # Lista de diccionarios con datos de cada período
        
        # Verificar licencia antes de mostrar la aplicación
        self.check_license()
        
    def extraer_datos_tabla(self, mes, anio):
        """Extrae los datos de la tabla de conceptos del HTML actual"""
        try:
            # Buscar la tabla de conceptos
            tabla = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "ctl00_PlaceContent_gvConceptos"))
            )
            
            # Extraer todas las filas de datos (excluyendo el header)
            filas = tabla.find_elements(By.CSS_SELECTOR, "tr.grilla_item, tr.grilla_aternateitem")
            
            conceptos_haberes = {}
            conceptos_deducciones = {}
            periodo = f"{mes:02d}/{anio}"
            
            for fila in filas:
                celdas = fila.find_elements(By.TAG_NAME, "td")
                if len(celdas) >= 5:
                    concepto = celdas[0].text.strip()
                    empresa = celdas[1].text.strip()
                    descripcion = celdas[2].text.strip()
                    haberes_text = celdas[3].text.strip()
                    deducciones_text = celdas[4].text.strip()
                    
                    # Crear clave única para el concepto
                    clave_concepto = f"{concepto}-{empresa}: {descripcion}"
                    
                    # Procesar haberes (si existe valor)
                    if haberes_text and haberes_text != "":
                        try:
                            valor_haber = float(haberes_text.replace(",", "."))
                            conceptos_haberes[clave_concepto] = valor_haber
                        except ValueError:
                            pass
                    
                    # Procesar deducciones (si existe valor)
                    if deducciones_text and deducciones_text != "":
                        try:
                            valor_deduccion = float(deducciones_text.replace(",", "."))
                            conceptos_deducciones[clave_concepto] = valor_deduccion
                        except ValueError:
                            pass
            
            self.console.log(f"Extraídos {len(conceptos_haberes)} haberes y {len(conceptos_deducciones)} deducciones del período {periodo}", "info")
            
            return {
                'periodo': periodo,
                'haberes': conceptos_haberes,
                'deducciones': conceptos_deducciones
            }
            
        except Exception as e:
            self.console.log(f"Error extrayendo datos de tabla: {e}", "error")
            return None

    def actualizar_excel(self, datos_periodo, ruta_excel):
        """Actualiza el archivo Excel con los datos del período - VERSIÓN MEJORADA"""
        try:
            # NUEVO: Almacenar datos para procesamiento posterior
            self.todos_los_datos.append(datos_periodo)
            
            periodo = datos_periodo['periodo']
            haberes = datos_periodo['haberes']
            deducciones = datos_periodo['deducciones']
            
            self.console.log(f"Actualizando Excel para período {periodo}...", "process")
            
            # Intentar cargar archivo existente o crear uno nuevo
            try:
                wb = load_workbook(ruta_excel)
            except FileNotFoundError:
                wb = Workbook()
                # Eliminar la hoja por defecto
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # Crear instancia del styler
            styler = ExcelStyler()
            
            # Crear/actualizar hoja de haberes
            if haberes:
                if 'Haberes' not in wb.sheetnames:
                    ws_haberes = wb.create_sheet('Haberes')
                else:
                    ws_haberes = wb['Haberes']
                
                self._actualizar_hoja_excel_corregida(ws_haberes, haberes, periodo, "HABERES", styler)
                self.console.log(f"✅ {len(haberes)} conceptos de haberes agregados", "success")
            
            # Crear/actualizar hoja de deducciones
            if deducciones:
                if 'Deducciones' not in wb.sheetnames:
                    ws_deducciones = wb.create_sheet('Deducciones')
                else:
                    ws_deducciones = wb['Deducciones']
                
                self._actualizar_hoja_excel_corregida(ws_deducciones, deducciones, periodo, "DEDUCCIONES", styler)
                self.console.log(f"✅ {len(deducciones)} conceptos de deducciones agregados", "success")
            
            # Guardar archivo
            wb.save(ruta_excel)
            self.console.log(f"📊 Excel actualizado exitosamente para período {periodo}", "success")
            
        except Exception as e:
            self.console.log(f"❌ Error actualizando Excel: {e}", "error")

    def crear_resumen_neto_completo(self, ruta_excel):
        """Crea la hoja de resumen neto con TODOS los datos recolectados"""
        try:
            self.console.log("🔄 Creando resumen neto completo...", "process")
            
            # Cargar workbook
            wb = load_workbook(ruta_excel)
            styler = ExcelStyler()
            
            # Recolectar TODOS los conceptos únicos de todos los períodos
            todos_haberes = set()
            todas_deducciones = set()
            
            for datos in self.todos_los_datos:
                todos_haberes.update(datos['haberes'].keys())
                todas_deducciones.update(datos['deducciones'].keys())
            
            # Ordenar conceptos
            haberes_ordenados = sorted(list(todos_haberes))
            deducciones_ordenadas = sorted(list(todas_deducciones))
            
            self.console.log(f"📊 Conceptos únicos encontrados: {len(haberes_ordenados)} haberes, {len(deducciones_ordenadas)} deducciones", "info")
            
            # Crear o limpiar hoja
            if 'Resumen Neto Detallado' in wb.sheetnames:
                ws_neto = wb['Resumen Neto Detallado']
                ws_neto.delete_rows(1, ws_neto.max_row)
            else:
                ws_neto = wb.create_sheet('Resumen Neto Detallado')
            
            # CREAR ESTRUCTURA COMPLETA
            columna_actual = 1
            
            # 1. Título principal
            title_cell = ws_neto.cell(row=1, column=1, value="💰 RESUMEN NETO DETALLADO - ANÁLISIS COMPLETO")
            title_cell.font = Font(name='Calibri', size=18, bold=True, color='FF1E40AF')
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 2. Subtitle con fecha
            subtitle = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Conceptos: H:{len(haberes_ordenados)} D:{len(deducciones_ordenadas)}"
            subtitle_cell = ws_neto.cell(row=2, column=1, value=subtitle)
            subtitle_cell.font = Font(name='Calibri', size=11, italic=True, color='FF6B7280')
            
            # 3. HEADERS
            # Columna PERÍODO
            period_header = ws_neto.cell(row=3, column=columna_actual, value="PERÍODO")
            self._apply_header_style_especial(period_header, styler, 'primary')
            ws_neto.column_dimensions[get_column_letter(columna_actual)].width = 12
            columna_actual += 1
            
            # SECCIÓN HABERES
            inicio_haberes = columna_actual
            if haberes_ordenados:
                # Header de sección
                seccion_haberes = ws_neto.cell(row=2, column=columna_actual, value="🟢 HABERES")
                seccion_haberes.font = Font(name='Calibri', size=14, bold=True, color='FF10B981')
                
                # Headers individuales
                for concepto in haberes_ordenados:
                    header_cell = ws_neto.cell(row=3, column=columna_actual, value=concepto)
                    self._apply_header_style_especial(header_cell, styler, 'haberes_bg')
                    ws_neto.column_dimensions[get_column_letter(columna_actual)].width = min(max(len(concepto) + 2, 15), 25)
                    columna_actual += 1
            
            # TOTAL HABERES
            total_haberes_col = columna_actual
            total_haberes_header = ws_neto.cell(row=3, column=columna_actual, value="TOTAL HABERES")
            self._apply_header_style_especial(total_haberes_header, styler, 'totales_bg')
            ws_neto.column_dimensions[get_column_letter(columna_actual)].width = 15
            columna_actual += 1
            
            # SECCIÓN DEDUCCIONES
            inicio_deducciones = columna_actual
            if deducciones_ordenadas:
                # Header de sección
                seccion_deducciones = ws_neto.cell(row=2, column=columna_actual, value="🔴 DEDUCCIONES")
                seccion_deducciones.font = Font(name='Calibri', size=14, bold=True, color='FFEF4444')
                
                # Headers individuales
                for concepto in deducciones_ordenadas:
                    header_cell = ws_neto.cell(row=3, column=columna_actual, value=concepto)
                    self._apply_header_style_especial(header_cell, styler, 'deducciones_bg')
                    ws_neto.column_dimensions[get_column_letter(columna_actual)].width = min(max(len(concepto) + 2, 15), 25)
                    columna_actual += 1
            
            # TOTAL DEDUCCIONES
            total_deducciones_col = columna_actual
            total_deducciones_header = ws_neto.cell(row=3, column=columna_actual, value="TOTAL DEDUCCIONES")
            self._apply_header_style_especial(total_deducciones_header, styler, 'totales_bg')
            ws_neto.column_dimensions[get_column_letter(columna_actual)].width = 18
            columna_actual += 1
            
            # NETO FINAL
            neto_col = columna_actual
            neto_header = ws_neto.cell(row=3, column=columna_actual, value="NETO (H-D)")
            self._apply_header_style_especial(neto_header, styler, 'primary')
            ws_neto.column_dimensions[get_column_letter(columna_actual)].width = 15
            
            # LLENAR DATOS
            fila_actual = 4
            for datos in self.todos_los_datos:
                periodo = datos['periodo']
                haberes = datos['haberes']
                deducciones = datos['deducciones']
                
                # Columna período
                period_cell = ws_neto.cell(row=fila_actual, column=1, value=periodo)
                self._apply_period_style(period_cell, styler)
                
                # Haberes individuales
                col_haber = inicio_haberes
                for concepto in haberes_ordenados:
                    valor = haberes.get(concepto, 0)
                    cell = ws_neto.cell(row=fila_actual, column=col_haber, value=valor if valor > 0 else None)
                    if valor > 0:
                        self._apply_data_style_especial(cell, valor, styler, 'haber')
                    col_haber += 1
                
                # Total haberes (fórmula)
                if haberes_ordenados:
                    inicio_col = get_column_letter(inicio_haberes)
                    fin_col = get_column_letter(inicio_haberes + len(haberes_ordenados) - 1)
                    formula = f"=SUM({inicio_col}{fila_actual}:{fin_col}{fila_actual})"
                else:
                    formula = 0
                
                total_haberes_cell = ws_neto.cell(row=fila_actual, column=total_haberes_col, value=formula)
                self._apply_data_style_especial(total_haberes_cell, None, styler, 'total_haber')
                
                # Deducciones individuales
                col_deduccion = inicio_deducciones
                for concepto in deducciones_ordenadas:
                    valor = deducciones.get(concepto, 0)
                    cell = ws_neto.cell(row=fila_actual, column=col_deduccion, value=valor if valor > 0 else None)
                    if valor > 0:
                        self._apply_data_style_especial(cell, valor, styler, 'deduccion')
                    col_deduccion += 1
                
                # Total deducciones (fórmula)
                if deducciones_ordenadas:
                    inicio_col = get_column_letter(inicio_deducciones)
                    fin_col = get_column_letter(inicio_deducciones + len(deducciones_ordenadas) - 1)
                    formula = f"=SUM({inicio_col}{fila_actual}:{fin_col}{fila_actual})"
                else:
                    formula = 0
                
                total_deducciones_cell = ws_neto.cell(row=fila_actual, column=total_deducciones_col, value=formula)
                self._apply_data_style_especial(total_deducciones_cell, None, styler, 'total_deduccion')
                
                # Neto final (fórmula)
                col_total_haberes = get_column_letter(total_haberes_col)
                col_total_deducciones = get_column_letter(total_deducciones_col)
                neto_formula = f"={col_total_haberes}{fila_actual}-{col_total_deducciones}{fila_actual}"
                
                neto_cell = ws_neto.cell(row=fila_actual, column=neto_col, value=neto_formula)
                self._apply_data_style_especial(neto_cell, None, styler, 'neto')
                
                # Aplicar formato de fila alternativa
                self._apply_alternate_row_formatting(ws_neto, fila_actual, styler)
                
                fila_actual += 1
            
            # Configuraciones finales
            ws_neto.freeze_panes = 'B4'
            if ws_neto.max_row > 3:
                end_col = get_column_letter(ws_neto.max_column)
                ws_neto.auto_filter.ref = f"A3:{end_col}{ws_neto.max_row}"
            
            # Ajustar dimensiones
            ws_neto.row_dimensions[1].height = 25
            ws_neto.row_dimensions[2].height = 20
            ws_neto.row_dimensions[3].height = 25
            
            # Guardar
            wb.save(ruta_excel)
            
            self.console.log("✅ Resumen neto detallado creado exitosamente", "success")
            self.console.log(f"📊 Procesados {len(self.todos_los_datos)} períodos con {len(haberes_ordenados) + len(deducciones_ordenadas)} conceptos únicos", "info")
            
        except Exception as e:
            self.console.log(f"❌ Error creando resumen neto completo: {e}", "error")

    def _actualizar_hoja_excel_corregida(self, worksheet, conceptos_dict, periodo, tipo_hoja, styler):
        """Actualiza una hoja específica del Excel - VERSIÓN CORREGIDA SIN MÚLTIPLES TOTALES"""
        try:
            # Configurar título de la hoja
            worksheet.title = tipo_hoja.title()
            
            # Verificar si necesita estructura inicial
            if worksheet.max_row <= 1:
                self._crear_estructura_hoja_simple(worksheet, tipo_hoja, styler)
                self.console.log(f"📋 Estructura inicial creada para hoja {tipo_hoja}", "info")
            
            # Obtener datos existentes
            periodos_existentes = {}
            conceptos_existentes = {}
            max_row = worksheet.max_row
            
            # IMPORTANTE: Verificar si ya existe columna TOTAL
            total_column = None
            max_col_datos = 1  # Empezar desde columna 1 (PERÍODO)
            
            for col in range(2, worksheet.max_column + 1):
                header_value = worksheet.cell(row=3, column=col).value
                if header_value == "TOTAL":
                    total_column = col
                    break
                elif header_value:  # Es una columna de concepto
                    max_col_datos = col
            
            # Leer períodos existentes desde la fila 4
            for row in range(4, max_row + 1):
                periodo_cell = worksheet.cell(row=row, column=1).value
                if periodo_cell:
                    periodos_existentes[str(periodo_cell)] = row
            
            # Leer conceptos existentes (fila 3, excluyendo TOTAL)
            for col in range(2, max_col_datos + 1):
                concepto_header = worksheet.cell(row=3, column=col).value
                if concepto_header and concepto_header != "TOTAL":
                    conceptos_existentes[concepto_header] = col
            
            # Agregar nuevos conceptos como columnas
            nueva_columna = max_col_datos + 1
            conceptos_agregados = 0
            
            for concepto in conceptos_dict.keys():
                if concepto not in conceptos_existentes:
                    # Crear header del concepto
                    cell = worksheet.cell(row=3, column=nueva_columna, value=concepto)
                    self._apply_header_style(cell, styler)
                    
                    # Ajustar ancho de columna
                    column_letter = get_column_letter(nueva_columna)
                    worksheet.column_dimensions[column_letter].width = max(15, min(len(concepto) + 2, 30))
                    
                    conceptos_existentes[concepto] = nueva_columna
                    nueva_columna += 1
                    conceptos_agregados += 1
            
            if conceptos_agregados > 0:
                self.console.log(f"➕ {conceptos_agregados} nuevos conceptos agregados", "info")
            
            # Determinar fila del período
            periodo_str = str(periodo)
            if periodo_str in periodos_existentes:
                fila_periodo = periodos_existentes[periodo_str]
                self.console.log(f"📅 Actualizando período existente {periodo} en fila {fila_periodo}", "info")
            else:
                fila_periodo = max_row + 1
                # Crear celda de período
                period_cell = worksheet.cell(row=fila_periodo, column=1, value=periodo)
                self._apply_period_style(period_cell, styler)
                self._apply_alternate_row_formatting(worksheet, fila_periodo, styler)
                self.console.log(f"📅 Nuevo período {periodo} agregado en fila {fila_periodo}", "info")
            
            # Agregar valores de conceptos
            valores_agregados = 0
            for concepto, valor in conceptos_dict.items():
                if concepto in conceptos_existentes:
                    columna_concepto = conceptos_existentes[concepto]
                    cell = worksheet.cell(row=fila_periodo, column=columna_concepto, value=valor)
                    self._apply_data_style(cell, valor, styler)
                    valores_agregados += 1
            
            self.console.log(f"💰 {valores_agregados} valores agregados para período {periodo}", "info")
            
            # Aplicar configuraciones finales solo una vez
            if worksheet.freeze_panes is None:
                worksheet.freeze_panes = 'B4'
            
            # Aplicar filtros automáticos
            if worksheet.max_row > 3:
                end_col = get_column_letter(worksheet.max_column)
                worksheet.auto_filter.ref = f"A3:{end_col}{worksheet.max_row}"
            
        except Exception as e:
            self.console.log(f"❌ Error actualizando hoja {tipo_hoja}: {e}", "error")

    def crear_columnas_total_finales(self, ruta_excel):
        """Crea las columnas TOTAL al final, después de procesar todos los períodos"""
        try:
            self.console.log("🔄 Creando columnas TOTAL finales...", "process")
            
            wb = load_workbook(ruta_excel)
            styler = ExcelStyler()
            
            # Procesar hojas de Haberes y Deducciones
            for sheet_name in ['Haberes', 'Deducciones']:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Verificar si ya existe columna TOTAL
                    total_exists = False
                    for col in range(2, ws.max_column + 1):
                        if ws.cell(row=3, column=col).value == "TOTAL":
                            total_exists = True
                            break
                    
                    if not total_exists:
                        # Crear columna TOTAL al final
                        total_col = ws.max_column + 1
                        self._crear_columna_total(ws, total_col, styler)
                        
                        # Actualizar fórmulas para todas las filas de datos
                        for row in range(4, ws.max_row + 1):
                            if ws.cell(row=row, column=1).value:  # Si hay período
                                self._actualizar_formula_total(ws, row, total_col)
                        
                        self.console.log(f"✅ Columna TOTAL creada para {sheet_name}", "success")
            
            wb.save(ruta_excel)
            self.console.log("✅ Columnas TOTAL finales creadas exitosamente", "success")
            
        except Exception as e:
            self.console.log(f"❌ Error creando columnas TOTAL: {e}", "error")

    # ===============================
    # FUNCIONES AUXILIARES CORREGIDAS
    # ===============================
    def _crear_estructura_hoja_simple(self, worksheet, tipo_hoja, styler):
        """Crea la estructura inicial de una hoja - VERSIÓN SIMPLE"""
        try:
            # Limpiar hoja completamente
            worksheet.delete_rows(1, worksheet.max_row)
            
            # Título principal (fila 1)
            title_cell = worksheet.cell(row=1, column=1, value=f"📊 {tipo_hoja}")
            title_cell.font = Font(name='Calibri', size=18, bold=True, color='FF1E40AF')
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Subtitle con fecha (fila 2)
            subtitle = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle)
            subtitle_cell.font = Font(name='Calibri', size=11, italic=True, color='FF6B7280')
            
            # Header de período (fila 3, columna 1)
            period_header = worksheet.cell(row=3, column=1, value="PERÍODO")
            period_header.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
            period_header.fill = PatternFill(
                start_color=styler.colors['primary'],
                end_color=styler.colors['primary'],
                fill_type='solid'
            )
            period_header.alignment = Alignment(horizontal='center', vertical='center')
            period_header.border = Border(
                left=Side(style='thick', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thick', color='FF000000'),
                bottom=Side(style='thick', color='FF000000')
            )
            
            # Ajustar dimensiones
            worksheet.column_dimensions['A'].width = 12
            worksheet.row_dimensions[1].height = 25
            worksheet.row_dimensions[2].height = 18
            worksheet.row_dimensions[3].height = 20
            
        except Exception as e:
            self.console.log(f"❌ Error creando estructura: {e}", "error")

    def _crear_columna_total(self, worksheet, columna_total, styler):
        """Crea la columna TOTAL en la posición especificada"""
        try:
            # Header TOTAL
            total_header = worksheet.cell(row=3, column=columna_total, value="TOTAL")
            total_header.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
            total_header.fill = PatternFill(
                start_color=styler.colors['secondary'],
                end_color=styler.colors['secondary'],
                fill_type='solid'
            )
            total_header.alignment = Alignment(horizontal='center', vertical='center')
            total_header.border = Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thick', color='FF000000'),
                top=Side(style='thick', color='FF000000'),
                bottom=Side(style='thick', color='FF000000')
            )
            
            # Ajustar ancho
            worksheet.column_dimensions[get_column_letter(columna_total)].width = 12
            
        except Exception as e:
            self.console.log(f"❌ Error creando columna TOTAL: {e}", "error")

    def _actualizar_formula_total(self, worksheet, fila, columna_total):
        """Actualiza la fórmula de total para una fila específica"""
        try:
            total_cell = worksheet.cell(row=fila, column=columna_total)
            
            # Solo actualizar si la celda está vacía o tiene una fórmula antigua
            if not total_cell.value or str(total_cell.value).startswith('='):
                start_col = get_column_letter(2)
                end_col = get_column_letter(columna_total - 1)
                total_cell.value = f"=SUM({start_col}{fila}:{end_col}{fila})"
                total_cell.number_format = '$#,##0.00'
                total_cell.font = Font(name='Calibri', size=11, bold=True, color='FF1E40AF')
                total_cell.alignment = Alignment(horizontal='right', vertical='center')
                total_cell.border = Border(
                    left=Side(style='thin', color='FFE5E7EB'),
                    right=Side(style='thick', color='FF1E40AF'),
                    top=Side(style='thin', color='FFE5E7EB'),
                    bottom=Side(style='thin', color='FFE5E7EB')
                )
            
        except Exception as e:
            self.console.log(f"❌ Error actualizando fórmula TOTAL: {e}", "error")

    def _apply_header_style_especial(self, cell, styler, tipo_color):
        """Aplica estilo de header especial según el tipo"""
        try:
            color_map = {
                'primary': styler.colors['primary'],
                'haberes_bg': styler.colors['haberes_bg'],
                'deducciones_bg': styler.colors['deducciones_bg'],
                'totales_bg': styler.colors['totales_bg']
            }
            
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(
                start_color=color_map.get(tipo_color, styler.colors['primary']),
                end_color=color_map.get(tipo_color, styler.colors['primary']),
                fill_type='solid'
            )
            cell.border = Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thick', color='FF000000')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except Exception as e:
            pass

    def _apply_data_style_especial(self, cell, valor, styler, tipo):
        """Aplica estilo de datos especial según el tipo"""
        try:
            cell.border = Border(
                left=Side(style='thin', color='FFE5E7EB'),
                right=Side(style='thin', color='FFE5E7EB'),
                top=Side(style='thin', color='FFE5E7EB'),
                bottom=Side(style='thin', color='FFE5E7EB')
            )
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '$#,##0.00'
            
            # Colores según tipo
            color_map = {
                'haber': 'FF10B981',           # Verde para haberes
                'deduccion': 'FFEF4444',       # Rojo para deducciones
                'total_haber': 'FF059669',     # Verde oscuro para total haberes
                'total_deduccion': 'FFDC2626', # Rojo oscuro para total deducciones
                'neto': 'FF1E40AF'             # Azul para neto
            }
            
            color = color_map.get(tipo, 'FF374151')
            cell.font = Font(name='Calibri', size=11, color=color, bold=(tipo.startswith('total') or tipo == 'neto'))
            
            # Borde especial para totales
            if tipo.startswith('total') or tipo == 'neto':
                cell.border = Border(
                    left=Side(style='thin', color='FFE5E7EB'),
                    right=Side(style='thick', color=color),
                    top=Side(style='thin', color='FFE5E7EB'),
                    bottom=Side(style='thin', color='FFE5E7EB')
                )
                
        except Exception as e:
            pass

    def _apply_header_style(self, cell, styler):
        """Aplica estilo de header a una celda"""
        try:
            cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(
                start_color=styler.colors['header_bg'],
                end_color=styler.colors['header_bg'],
                fill_type='solid'
            )
            cell.border = Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thick', color='FF000000')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except Exception as e:
            pass  # Ignorar errores de estilo

    def _apply_period_style(self, cell, styler):
        """Aplica estilo de período a una celda"""
        try:
            cell.font = Font(name='Calibri', size=11, bold=True, color='FF1E40AF')
            cell.fill = PatternFill(
                start_color=styler.colors['light'],
                end_color=styler.colors['light'],
                fill_type='solid'
            )
            cell.border = Border(
                left=Side(style='thick', color='FF1E40AF'),
                right=Side(style='thin', color='FFE5E7EB'),
                top=Side(style='thin', color='FFE5E7EB'),
                bottom=Side(style='thin', color='FFE5E7EB')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')
        except Exception as e:
            pass

    def _apply_data_style(self, cell, valor, styler):
        """Aplica estilo de datos a una celda"""
        try:
            cell.border = Border(
                left=Side(style='thin', color='FFE5E7EB'),
                right=Side(style='thin', color='FFE5E7EB'),
                top=Side(style='thin', color='FFE5E7EB'),
                bottom=Side(style='thin', color='FFE5E7EB')
            )
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '$#,##0.00'
            
            # Formato condicional para valores
            if valor > 0:
                cell.font = Font(name='Calibri', size=11, color='FF059669')  # Verde
            elif valor < 0:
                cell.font = Font(name='Calibri', size=11, color='FFEF4444')  # Rojo
            else:
                cell.font = Font(name='Calibri', size=11, color='FF6B7280')  # Gris
        except Exception as e:
            pass

    def _apply_alternate_row_formatting(self, worksheet, fila, styler):
        """Aplica formato de fila alternativa"""
        try:
            if (fila - 4) % 2 == 1:  # Filas impares
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=fila, column=col)
                    if cell.fill.start_color.index in ['00000000', '00FFFFFF']:  # Sin color o blanco
                        cell.fill = PatternFill(
                            start_color=styler.colors['alt_row'],
                            end_color=styler.colors['alt_row'],
                            fill_type='solid'
                        )
        except Exception as e:
            pass

    def abrir_excel(self):
        """Abre el archivo Excel de análisis"""
        ruta_excel = os.path.join(self.carpeta_entry.get(), "analisis_recibos.xlsx")
        if os.path.exists(ruta_excel):
            try:
                if platform.system() == "Windows":
                    os.startfile(ruta_excel)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.run(["open", ruta_excel])
                else:  # Linux
                    subprocess.run(["xdg-open", ruta_excel])
                self.console.log(f"Abriendo Excel: analisis_recibos.xlsx", "success")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el Excel:\n{e}")
                self.console.log(f"Error abriendo Excel: {e}", "error")
        else:
            messagebox.showwarning("Archivo no encontrado", "No se encontró el archivo Excel de análisis")
            self.console.log("Excel de análisis no encontrado", "warning")
    
    def check_license(self):
        """Verifica la licencia antes de iniciar la aplicación"""
        # Mostrar diálogo de licencia
        license_dialog = LicenseDialog(self.root, self.license_manager)
        self.root.wait_window(license_dialog)
        
        # Verificar si se cerró con licencia válida
        if not license_dialog.license_valid:
            messagebox.showerror("Acceso Denegado", "No tienes una licencia válida.\nLa aplicación se cerrará.")
            self.root.quit()
            return
        
        # Si llegamos aquí, la licencia es válida
        self.setup_ui()
        self.start_animations()
        
        # Mostrar información de licencia en consola
        self.console.log(f"Licencia verificada para máquina: {self.license_manager.machine_id[:16]}...", "success")
        self.console.log("Sistema autorizado y listo para usar", "info")
    
    def setup_ui(self):
        # Frame principal con diseño de dos columnas
        main_container = ctk.CTkFrame(self.root, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Columna izquierda - Configuración
        left_panel = ctk.CTkFrame(main_container, width=500)
        left_panel.pack(side="left", fill="both", expand=False, padx=(0, 5))
        left_panel.pack_propagate(False)
        
        # Columna derecha - Consola y progreso
        right_panel = ctk.CTkFrame(main_container)
        right_panel.pack(side="right", fill="both", expand=True, padx=(5, 0))
        
        self.setup_left_panel(left_panel)
        self.setup_right_panel(right_panel)
    
    def setup_left_panel(self, parent):
        # Título principal con estilo corporativo
        title_frame = ctk.CTkFrame(parent, height=80, fg_color=["#1E3A8A", "#0F172A"])
        title_frame.pack(fill="x", padx=15, pady=(15, 20))
        title_frame.pack_propagate(False)

        title_label = ctk.CTkLabel(
            title_frame,
            text="⚖️ Sistema Legal de Recibos",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="white"
        )
        title_label.pack(expand=True)

        subtitle_label = ctk.CTkLabel(
            title_frame,
            text="Gestión Profesional para Estudios Jurídicos",
            font=ctk.CTkFont(size=12),
            text_color="#D1D5DB"
        )
        subtitle_label.pack()

        # Interruptor de tema para cambiar entre modo oscuro y claro
        self.theme_switch = ctk.CTkSwitch(
            title_frame,
            text="Modo oscuro",
            command=self.toggle_theme
        )
        self.theme_switch.select()
        self.theme_switch.place(relx=1, x=-10, y=10, anchor="ne")

        # Scrollable frame para configuración
        config_scroll = ctk.CTkScrollableFrame(parent, height=500)
        config_scroll.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # Sección de credenciales
        self.create_section(config_scroll, "🔐 CREDENCIALES DE ACCESO", [
            ("Usuario (CUIL):", "usuario_entry", "27-12345678-9", "27-03321528-8"),
            ("Contraseña:", "clave_entry", "Tu contraseña", "Belgrano1188", True),
            ("Número de Beneficio:", "beneficio_entry", "12345678901", "15584832490")
        ])
        
        # Sección de fechas
        self.create_date_section(config_scroll)
        
        # Sección de carpeta
        self.create_folder_section(config_scroll)
        
        # Botones de control
        self.create_control_buttons(config_scroll)
        
        # Información de licencia
        self.create_license_info(config_scroll)
    
    def create_license_info(self, parent):
        """Muestra información de la licencia"""
        license_frame = ctk.CTkFrame(parent, fg_color=["#F0FDF4", "#1F2937"])
        license_frame.pack(fill="x", pady=(0, 15))
        
        title_label = ctk.CTkLabel(
            license_frame,
            text="🔐 INFORMACIÓN DE LICENCIA",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=["#059669", "#10B981"]
        )
        title_label.pack(pady=(15, 10))
        
        info_label = ctk.CTkLabel(
            license_frame,
            text=f"✅ Máquina autorizada\n🆔 ID: {self.license_manager.machine_id[:32]}...",
            font=ctk.CTkFont(size=12),
            text_color=["#059669", "#10B981"]
        )
        info_label.pack(pady=(0, 15))

    def toggle_theme(self):
        """Alterna entre modo claro y oscuro"""
        if self.theme_switch.get():
            ctk.set_appearance_mode("dark")
            self.theme_switch.configure(text="Modo oscuro")
        else:
            ctk.set_appearance_mode("light")
            self.theme_switch.configure(text="Modo claro")

    def create_section(self, parent, title, fields):
        # Frame de sección con estilo
        section_frame = ctk.CTkFrame(parent, fg_color=["#F8FAFC", "#1E293B"])
        section_frame.pack(fill="x", pady=(0, 15))
        
        # Título de sección
        title_label = ctk.CTkLabel(
            section_frame,
            text=title,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=["#1E3A8A", "#93C5FD"]
        )
        title_label.pack(pady=(15, 10))
        
        # Campos
        for field in fields:
            field_frame = ctk.CTkFrame(section_frame, fg_color="transparent")
            field_frame.pack(fill="x", padx=20, pady=5)
            
            label = ctk.CTkLabel(
                field_frame,
                text=field[0],
                font=ctk.CTkFont(size=12, weight="bold"),
                anchor="w"
            )
            label.pack(anchor="w", pady=(5, 2))
            
            show_char = "*" if len(field) > 4 and field[4] else None
            entry = ctk.CTkEntry(
                field_frame,
                placeholder_text=field[2],
                height=35,
                font=ctk.CTkFont(size=12),
                show=show_char
            )
            entry.pack(fill="x", pady=(0, 10))
            
            if len(field) > 3:
                entry.insert(0, field[3])
            
            setattr(self, field[1], entry)
        
        # Espaciado final
        ctk.CTkLabel(section_frame, text="", height=10).pack()
    
    def create_date_section(self, parent):
        section_frame = ctk.CTkFrame(parent, fg_color=["#F8FAFC", "#1E293B"])
        section_frame.pack(fill="x", pady=(0, 15))
        
        title_label = ctk.CTkLabel(
            section_frame,
            text="📅 RANGO DE FECHAS",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=["#1E3A8A", "#93C5FD"]
        )
        title_label.pack(pady=(15, 10))
        
        # Fecha inicial
        fecha_inicial_frame = ctk.CTkFrame(section_frame, fg_color="transparent")
        fecha_inicial_frame.pack(fill="x", padx=20, pady=5)
        
        ctk.CTkLabel(
            fecha_inicial_frame,
            text="Desde:",
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(anchor="w", pady=(5, 2))
        
        fecha_inicial_inputs = ctk.CTkFrame(fecha_inicial_frame, fg_color="transparent")
        fecha_inicial_inputs.pack(fill="x")
        
        self.mes_inicial_entry = ctk.CTkEntry(
            fecha_inicial_inputs,
            placeholder_text="MM",
            width=80,
            height=35,
            font=ctk.CTkFont(size=12)
        )
        self.mes_inicial_entry.pack(side="left", padx=(0, 10))
        self.mes_inicial_entry.insert(0, "6")
        
        ctk.CTkLabel(fecha_inicial_inputs, text="/", font=ctk.CTkFont(size=16, weight="bold")).pack(side="left", padx=5)
        
        self.anio_inicial_entry = ctk.CTkEntry(
            fecha_inicial_inputs,
            placeholder_text="AAAA",
            width=100,
            height=35,
            font=ctk.CTkFont(size=12)
        )
        self.anio_inicial_entry.pack(side="left", padx=(10, 0))
        self.anio_inicial_entry.insert(0, "2024")
        
        # Fecha final
        fecha_final_frame = ctk.CTkFrame(section_frame, fg_color="transparent")
        fecha_final_frame.pack(fill="x", padx=20, pady=5)
        
        ctk.CTkLabel(
            fecha_final_frame,
            text="Hasta:",
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(anchor="w", pady=(5, 2))
        
        fecha_final_inputs = ctk.CTkFrame(fecha_final_frame, fg_color="transparent")
        fecha_final_inputs.pack(fill="x")
        
        self.mes_final_entry = ctk.CTkEntry(
            fecha_final_inputs,
            placeholder_text="MM",
            width=80,
            height=35,
            font=ctk.CTkFont(size=12)
        )
        self.mes_final_entry.pack(side="left", padx=(0, 10))
        self.mes_final_entry.insert(0, "11")
        
        ctk.CTkLabel(fecha_final_inputs, text="/", font=ctk.CTkFont(size=16, weight="bold")).pack(side="left", padx=5)
        
        self.anio_final_entry = ctk.CTkEntry(
            fecha_final_inputs,
            placeholder_text="AAAA",
            width=100,
            height=35,
            font=ctk.CTkFont(size=12)
        )
        self.anio_final_entry.pack(side="left", padx=(10, 0))
        self.anio_final_entry.insert(0, "2024")
        
        ctk.CTkLabel(section_frame, text="", height=10).pack()
    
    def create_folder_section(self, parent):
        section_frame = ctk.CTkFrame(parent, fg_color=["#F8FAFC", "#1E293B"])
        section_frame.pack(fill="x", pady=(0, 15))
        
        title_label = ctk.CTkLabel(
            section_frame,
            text="📁 CARPETA DE DESCARGA",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=["#1E3A8A", "#93C5FD"]
        )
        title_label.pack(pady=(15, 10))
        
        folder_frame = ctk.CTkFrame(section_frame, fg_color="transparent")
        folder_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        self.carpeta_entry = ctk.CTkEntry(
            folder_frame,
            placeholder_text="Selecciona la carpeta de descarga",
            height=35,
            font=ctk.CTkFont(size=12)
        )
        self.carpeta_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.carpeta_entry.insert(0, os.path.join(os.getcwd(), "recibos_descargados"))
        
        folder_btn = ctk.CTkButton(
            folder_frame,
            text="📂",
            width=50,
            height=35,
            command=self.seleccionar_carpeta,
            font=ctk.CTkFont(size=16)
        )
        folder_btn.pack(side="right")
    
    def create_control_buttons(self, parent):
        # Frame para botones con estilo especial
        buttons_frame = ctk.CTkFrame(parent, fg_color="transparent")
        buttons_frame.pack(fill="x", pady=20)
        
        # Botón principal de inicio/parada
        self.start_btn = ctk.CTkButton(
            buttons_frame,
            text="🚀 INICIAR DESCARGA",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=50,
            fg_color=["#10B981", "#059669"],
            hover_color=["#059669", "#047857"],
            command=self.toggle_descarga
        )
        self.start_btn.pack(fill="x", pady=(0, 10))
        
        # Botón para abrir PDF - INICIALMENTE DESHABILITADO
        self.pdf_btn = ctk.CTkButton(
            buttons_frame,
            text="📄 PDF NO DISPONIBLE",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=45,
            fg_color=["#6B7280", "#4B5563"],  # Gris deshabilitado
            hover_color=["#6B7280", "#4B5563"],
            command=self.abrir_pdf,
            state="disabled"
        )
        self.pdf_btn.pack(fill="x", pady=(0, 10))
        
        # Botón para abrir Excel - INICIALMENTE DESHABILITADO
        self.excel_btn = ctk.CTkButton(
            buttons_frame,
            text="📊 EXCEL NO DISPONIBLE",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=45,
            fg_color=["#6B7280", "#4B5563"],  # Gris deshabilitado
            hover_color=["#6B7280", "#4B5563"],
            command=self.abrir_excel,
            state="disabled"
        )
        self.excel_btn.pack(fill="x")
    
    def setup_right_panel(self, parent):
        # Título del panel derecho
        right_title = ctk.CTkLabel(
            parent,
            text="Panel de Monitoreo",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=["#1E3A8A", "#93C5FD"]
        )
        right_title.pack(pady=(15, 10))
        
        # Frame de estadísticas
        stats_frame = ctk.CTkFrame(parent, height=80, fg_color=["#EFF6FF", "#1E293B"])
        stats_frame.pack(fill="x", padx=15, pady=(0, 15))
        stats_frame.pack_propagate(False)
        
        stats_container = ctk.CTkFrame(stats_frame, fg_color="transparent")
        stats_container.pack(expand=True, fill="both", padx=10, pady=10)
        
        # Estadísticas en tiempo real
        self.stats_labels = {}
        stats_data = [
            ("Estado:", "🔴 Detenido", "status"),
            ("Progreso:", "0%", "progress"),
            ("Archivos:", "0 PDFs", "files")
        ]
        
        for i, (label, value, key) in enumerate(stats_data):
            stat_frame = ctk.CTkFrame(stats_container, fg_color="transparent")
            stat_frame.pack(side="left", fill="both", expand=True, padx=5)
            
            ctk.CTkLabel(
                stat_frame,
                text=label,
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=["#6B7280", "#9CA3AF"]
            ).pack()
            
            value_label = ctk.CTkLabel(
                stat_frame,
                text=value,
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=["#1F2937", "#F9FAFB"]
            )
            value_label.pack()
            self.stats_labels[key] = value_label
        
        # Barra de progreso mejorada
        progress_frame = ctk.CTkFrame(parent, fg_color=["#F8FAFC", "#1E293B"])
        progress_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        ctk.CTkLabel(
            progress_frame,
            text="⚡ PROGRESO DE DESCARGA",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=["#1E3A8A", "#93C5FD"]
        ).pack(pady=(15, 5))
        
        self.progress_bar = ctk.CTkProgressBar(
            progress_frame,
            height=20,
            progress_color=["#10B981", "#059669"]
        )
        self.progress_bar.pack(fill="x", padx=20, pady=(0, 10))
        self.progress_bar.set(0)
        
        self.progress_label = ctk.CTkLabel(
            progress_frame,
            text="0 / 0 recibos procesados",
            font=ctk.CTkFont(size=12),
            text_color=["#6B7280", "#9CA3AF"]
        )
        self.progress_label.pack(pady=(0, 15))
        
        # Consola profesional
        console_frame = ctk.CTkFrame(parent, fg_color=["#F8FAFC", "#1E293B"])
        console_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        console_header = ctk.CTkFrame(console_frame, height=40, fg_color=["#1F2937", "#111827"])
        console_header.pack(fill="x", padx=2, pady=(2, 0))
        console_header.pack_propagate(False)
        
        ctk.CTkLabel(
            console_header,
            text="Consola del Sistema",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#D1D5DB"
        ).pack(side="left", padx=15, pady=10)
        
        # Botón para limpiar consola
        clear_btn = ctk.CTkButton(
            console_header,
            text="🗑️",
            width=30,
            height=25,
            font=ctk.CTkFont(size=12),
            fg_color="transparent",
            hover_color=["#374151", "#4B5563"],
            command=self.clear_console
        )
        clear_btn.pack(side="right", padx=15, pady=7)
        
        # Widget de consola personalizado
        self.console = ConsoleWidget(
            console_frame,
            height=300,
            wrap="word"
        )
        self.console.pack(fill="both", expand=True, padx=2, pady=(0, 2))
        
        # Mensaje inicial
        self.console.log("Sistema iniciado correctamente", "success")
        self.console.log("Esperando configuración del usuario...", "info")
    
    def clear_console(self):
        self.console.delete("1.0", "end")
        self.console.log("Consola limpiada", "info")
    
    def start_animations(self):
        """Inicia animaciones sutiles para la interfaz"""
        self.animation_running = True
        self.animate_title()
    
    def animate_title(self):
        """Animación sutil del título"""
        if not self.animation_running:
            return
        
        # Programar siguiente animación
        self.root.after(3000, self.animate_title)
    
    def seleccionar_carpeta(self):
        carpeta = filedialog.askdirectory()
        if carpeta:
            self.carpeta_entry.delete(0, tk.END)
            self.carpeta_entry.insert(0, carpeta)
            self.console.log(f"Carpeta seleccionada: {carpeta}", "info")
    
    def update_stats(self, status=None, progress=None, files=None):
        """Actualiza las estadísticas en tiempo real"""
        if status:
            self.stats_labels["status"].configure(text=status)
        if progress is not None:
            self.stats_labels["progress"].configure(text=f"{int(progress*100)}%")
        if files is not None:
            self.stats_labels["files"].configure(text=f"{files} PDFs")
    
    def iluminar_boton_pdf(self):
        """Ilumina el botón PDF cuando está disponible"""
        self.pdf_btn.configure(
            state="normal",
            fg_color=["#10B981", "#059669"],  # Verde brillante
            hover_color=["#059669", "#047857"],
            text="📄 ✨ ABRIR PDF UNIFICADO ✨"
        )
        self.console.log("✨ Botón PDF iluminado - ¡Archivo listo!", "success")

    def iluminar_boton_excel(self):
        """Ilumina el botón Excel cuando está disponible"""
        self.excel_btn.configure(
            state="normal",
            fg_color=["#F59E0B", "#D97706"],  # Naranja brillante
            hover_color=["#D97706", "#B45309"],
            text="📊 ✨ ABRIR ANÁLISIS EXCEL ✨"
        )
        self.console.log("✨ Botón Excel iluminado - ¡Análisis listo!", "success")
        
    def toggle_descarga(self):
        # Verificar licencia antes de iniciar descarga
        valid, message = self.license_manager.check_license()
        if not valid:
            messagebox.showerror("Licencia Inválida", f"No se puede iniciar la descarga:\n{message}")
            return
        
        if not self.is_running:
            self.iniciar_descarga()
        else:
            self.detener_descarga()
    
    def limpiar_pdfs_individuales(self, carpeta):
        """Elimina todos los PDFs excepto el unificado"""
        try:
            lista_pdfs = glob.glob(os.path.join(carpeta, "*.pdf"))
            eliminados = 0
            for pdf in lista_pdfs:
                if not pdf.endswith("todos_los_recibos.pdf"):
                    os.remove(pdf)
                    eliminados += 1
            if eliminados > 0:
                self.console.log(f"Limpieza completada: {eliminados} archivos temporales eliminados", "success")
        except Exception as e:
            self.console.log(f"Error en limpieza de archivos: {e}", "warning")

    def limpiar_carpeta_completa(self, carpeta):
        """Elimina todos los archivos de la carpeta al iniciar nueva descarga"""
        try:
            if os.path.exists(carpeta):
                archivos_eliminados = 0
                # Eliminar todos los archivos en la carpeta
                for archivo in os.listdir(carpeta):
                    ruta_archivo = os.path.join(carpeta, archivo)
                    if os.path.isfile(ruta_archivo):
                        os.remove(ruta_archivo)
                        archivos_eliminados += 1
                
                if archivos_eliminados > 0:
                    self.console.log(f"🗑️ Carpeta limpiada: {archivos_eliminados} archivos eliminados", "success")
                else:
                    self.console.log("📁 Carpeta ya estaba vacía", "info")
            else:
                self.console.log("📁 Carpeta no existe, se creará automáticamente", "info")
        except Exception as e:
            self.console.log(f"❌ Error limpiando carpeta: {e}", "error")

    def iniciar_descarga(self):
        # Validar campos
        if not all([
            self.usuario_entry.get(),
            self.clave_entry.get(),
            self.beneficio_entry.get(),
            self.mes_inicial_entry.get(),
            self.anio_inicial_entry.get(),
            self.mes_final_entry.get(),
            self.anio_final_entry.get()
        ]):
            messagebox.showerror("Error", "Por favor completa todos los campos")
            return
        
        # NUEVO: Limpiar carpeta completa al iniciar
        carpeta_descargas = self.carpeta_entry.get()
        self.limpiar_carpeta_completa(carpeta_descargas)
        
        # NUEVO: Limpiar datos anteriores
        self.todos_los_datos = []
        
        self.is_running = True
        self.start_btn.configure(
            text="⏹️ DETENER PROCESO",
            fg_color=["#EF4444", "#DC2626"],
            hover_color=["#DC2626", "#B91C1C"]
        )
        
        # NUEVO: Resetear botones a estado deshabilitado
        self.pdf_btn.configure(
            state="disabled",
            fg_color=["#6B7280", "#4B5563"],  # Gris deshabilitado
            hover_color=["#6B7280", "#4B5563"],
            text="📄 PDF NO DISPONIBLE"
        )
        self.excel_btn.configure(
            state="disabled", 
            fg_color=["#6B7280", "#4B5563"],  # Gris deshabilitado
            hover_color=["#6B7280", "#4B5563"],
            text="📊 EXCEL NO DISPONIBLE"
        )
        
        self.progress_bar.set(0)
        self.update_stats(status="🟡 Iniciando...", progress=0, files=0)
        
        # Ejecutar en hilo separado
        thread = threading.Thread(target=self.proceso_descarga)
        thread.daemon = True
        thread.start()
        
    def detener_descarga(self):
        self.is_running = False
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.start_btn.configure(
            text="🚀 INICIAR DESCARGA",
            fg_color=["#10B981", "#059669"],
            hover_color=["#059669", "#047857"]
        )
        self.update_stats(status="🔴 Detenido")
        self.console.log("Proceso detenido por el usuario", "warning")
    
    def proceso_descarga(self):
        try:
            # Obtener valores de la interfaz
            usuario = self.usuario_entry.get()
            clave = self.clave_entry.get()
            beneficio = self.beneficio_entry.get()
            mes_inicial = int(self.mes_inicial_entry.get())
            anio_inicial = int(self.anio_inicial_entry.get())
            mes_final = int(self.mes_final_entry.get())
            anio_final = int(self.anio_final_entry.get())
            carpeta_descargas = self.carpeta_entry.get()
            
            self.update_stats(status="🟡 Configurando...")
            
            # Crear carpeta si no existe
            if not os.path.exists(carpeta_descargas):
                os.makedirs(carpeta_descargas)
                self.console.log(f"Carpeta creada: {carpeta_descargas}", "success")
            
            # Configurar Chrome
            opciones = Options()
            opciones.add_argument("--start-maximized")
            prefs = {
                "download.default_directory": carpeta_descargas,
                "download.prompt_for_download": False,
                "plugins.always_open_pdf_externally": True
            }
            opciones.add_experimental_option("prefs", prefs)
            
            self.driver = webdriver.Chrome(options=opciones)
            self.console.log("Navegador Chrome iniciado correctamente", "success")
            self.update_stats(status="🟡 Conectando...")
            
            # Proceso de login
            url_login = "https://servicioscorp.anses.gob.ar/clavelogon/logon.aspx?system=miansesv2"
            self.driver.get(url_login)
            time.sleep(3)
            
            self.driver.find_element(By.ID, "Usuario").send_keys(usuario)
            self.driver.find_element(By.ID, "Clave").send_keys(clave)
            
            self.console.log("Credenciales ingresadas. Esperando resolución de CAPTCHA...", "process")
            self.update_stats(status="🟡 CAPTCHA...")
            
            try:
                WebDriverWait(self.driver, 30).until(
                    EC.element_to_be_clickable((By.ID, "Ingresar"))
                )
                boton = self.driver.find_element(By.ID, "Ingresar")
                boton.click()
                self.console.log("Intentando acceder al sistema...", "process")
                time.sleep(5)
            except Exception as e:
                self.console.log("Tiempo agotado. CAPTCHA no resuelto", "error")
                return
            
            self.update_stats(status="🟢 Conectado")
            
            # Cerrar notificación
            try:
                self.console.log("Cerrando notificaciones emergentes...", "process")
                boton_cerrar = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[@class='btn-close' and @aria-label='Cerrar']"))
                )
                boton_cerrar.click()
                time.sleep(1)
            except Exception as e:
                self.console.log("No se encontraron notificaciones para cerrar", "info")
            
            # Navegar a jubilaciones
            try:
                self.console.log("Navegando a 'Jubilaciones y pensiones'...", "process")
                WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//span[text()='Jubilaciones y pensiones']"))
                ).click()
                time.sleep(2)
            except Exception as e:
                self.console.log(f"Error navegando a jubilaciones: {e}", "error")
                return
            
            # Consultar recibos
            try:
                self.console.log("Accediendo a 'Consultar recibos de haberes'...", "process")
                WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(@data-href, '10603')]"))
                ).click()
                time.sleep(5)
            except Exception as e:
                self.console.log(f"Error accediendo a recibos: {e}", "error")
                return
            
            # Cambiar al iframe
            try:
                self.console.log("Configurando interfaz de recibos...", "process")
                iframe = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "iframe"))
                )
                self.driver.switch_to.frame(iframe)
                self.console.log("Interfaz configurada correctamente", "success")
            except Exception as e:
                self.console.log(f"Error configurando interfaz: {e}", "error")
                return
            
            # Calcular total de meses
            total_meses = self.calcular_total_meses(mes_inicial, anio_inicial, mes_final, anio_final)
            mes_actual = 0
            archivos_procesados = 0
            
            self.console.log(f"Iniciando descarga de {total_meses} recibos...", "success")
            self.update_stats(status="🟢 Descargando", files=0)
            
            # Inicializar merger y ruta del Excel
            pdf_merger = PdfMerger()
            self.pdf_final_path = os.path.join(carpeta_descargas, "todos_los_recibos.pdf")
            ruta_excel = os.path.join(carpeta_descargas, "analisis_recibos.xlsx")
            
            # Bucle de descarga
            mes = mes_inicial
            anio = anio_inicial
            
            while (anio < anio_final) or (anio == anio_final and mes <= mes_final):
                if not self.is_running:
                    break
                
                mes_actual += 1
                progreso = mes_actual / total_meses
                self.progress_bar.set(progreso)
                self.progress_label.configure(text=f"{mes_actual} / {total_meses} recibos procesados")
                self.update_stats(progress=progreso)
                
                self.console.log(f"Procesando recibo {mes:02d}/{anio} ({mes_actual}/{total_meses})", "process")
                
                # Seleccionar beneficio
                try:
                    select_benef = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.ID, "ctl00_PlaceContent_ddl_Beneficios"))
                    )
                    for option in select_benef.find_elements(By.TAG_NAME, "option"):
                        if option.get_attribute("value") == beneficio:
                            option.click()
                            break
                except Exception as e:
                    self.console.log(f"Error seleccionando beneficio: {e}", "error")
                    break
                
                # Ingresar mes y año
                try:
                    mes_input = self.driver.find_element(By.ID, "ctl00_PlaceContent_txtMes")
                    anio_input = self.driver.find_element(By.ID, "ctl00_PlaceContent_txtAnio")
                    mes_input.clear()
                    mes_input.send_keys(str(mes))
                    anio_input.clear()
                    anio_input.send_keys(str(anio))
                except Exception as e:
                    self.console.log(f"Error ingresando fecha: {e}", "error")
                    break
                
                # Consultar
                try:
                    self.driver.find_element(By.ID, "ctl00_PlaceContent_btnConsultar").click()
                    time.sleep(3)
                except Exception as e:
                    self.console.log(f"Error consultando recibo: {e}", "error")
                    break
                
                # ===== NUEVA FUNCIONALIDAD: EXTRAER DATOS PARA EXCEL =====
                datos_tabla = self.extraer_datos_tabla(mes, anio)
                if datos_tabla:
                    self.actualizar_excel(datos_tabla, ruta_excel)
                    self.console.log(f"Datos del período {mes:02d}/{anio} agregados al Excel", "success")
                # ===== FIN NUEVA FUNCIONALIDAD =====
                
                # Descargar PDF
                try:
                    self.console.log(f"Descargando PDF {mes:02d}/{anio}...", "process")
                    imprimir_btn = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.ID, "ctl00_PlaceContent_btn_imprimir"))
                    )
                    imprimir_btn.click()
                    time.sleep(3)
                    
                    # Buscar PDF más reciente
                    lista_pdfs = glob.glob(os.path.join(carpeta_descargas, "*.pdf"))
                    # Filtrar el archivo final si ya existe
                    lista_pdfs = [pdf for pdf in lista_pdfs if not pdf.endswith("todos_los_recibos.pdf")]
                    lista_pdfs.sort(key=os.path.getmtime, reverse=True)
                    
                    if lista_pdfs:
                        pdf_mas_reciente = lista_pdfs[0]
                        if self.esperar_descarga_completa(pdf_mas_reciente):
                            self.console.log(f"Uniendo archivo: {os.path.basename(pdf_mas_reciente)}", "success")
                            pdf_merger.append(pdf_mas_reciente)
                            # Eliminar PDF individual inmediatamente
                            os.remove(pdf_mas_reciente)
                            archivos_procesados += 1
                            self.update_stats(files=archivos_procesados)
                            self.console.log(f"Archivo procesado y eliminado: {os.path.basename(pdf_mas_reciente)}", "info")
                        else:
                            self.console.log("No se pudo procesar el PDF descargado", "warning")
                
                except Exception as e:
                    self.console.log(f"Error descargando PDF: {e}", "error")
                
                # Volver atrás
                try:
                    self.driver.back()
                    time.sleep(3)
                    iframe = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, "iframe"))
                    )
                    self.driver.switch_to.frame(iframe)
                except Exception as e:
                    self.console.log(f"Error navegando hacia atrás: {e}", "error")
                    break
                
                # Incrementar mes
                mes += 1
                if mes > 12:
                    mes = 1
                    anio += 1
            
            # Guardar PDF final
            if self.is_running:
                try:
                    self.console.log("Generando PDF unificado final...", "process")
                    self.update_stats(status="🟡 Finalizando...")
                    
                    # Eliminar archivo final anterior si existe
                    if os.path.exists(self.pdf_final_path):
                        os.remove(self.pdf_final_path)
                    
                    pdf_merger.write(self.pdf_final_path)
                    pdf_merger.close()
                    
                    # Limpiar cualquier PDF individual restante
                    self.limpiar_pdfs_individuales(carpeta_descargas)
                    
                    self.console.log(f"PDF unificado creado exitosamente: {self.pdf_final_path}", "success")
                    self.iluminar_boton_pdf()

                    # NUEVO: Crear resumen neto completo al final
                    if os.path.exists(ruta_excel) and self.todos_los_datos:
                        self.crear_resumen_neto_completo(ruta_excel)
                        self.crear_columnas_total_finales(ruta_excel)
                        self.iluminar_boton_excel()
                        self.console.log(f"📊 Análisis Excel completo creado: analisis_recibos.xlsx", "success")
                    
                    self.progress_bar.set(1.0)
                    self.progress_label.configure(text=f"¡Completado! {total_meses} recibos procesados")
                    self.update_stats(status="✅ Completado", progress=1.0)
                    
                    messagebox.showinfo(
                        "¡Descarga Completada!",
                        f"✅ Proceso finalizado exitosamente\n\n"
                        f"📊 Recibos procesados: {archivos_procesados}\n"
                        f"📄 PDF unificado: todos_los_recibos.pdf\n"
                        f"📈 Análisis Excel: analisis_recibos.xlsx\n"
                        f"📁 Ubicación: {carpeta_descargas}"
                    )
                
                except Exception as e:
                    self.console.log(f"Error creando PDF final: {e}", "error")
        
        except Exception as e:
            self.console.log(f"Error crítico del sistema: {e}", "error")
            messagebox.showerror("Error Crítico", f"Error durante el proceso:\n{e}")
        
        finally:
            self.is_running = False
            self.start_btn.configure(
                text="🚀 INICIAR DESCARGA",
                fg_color=["#10B981", "#059669"],
                hover_color=["#059669", "#047857"]
            )
            if self.driver:
                try:
                    self.driver.quit()
                    self.console.log("Navegador cerrado correctamente", "info")
                except:
                    pass
    
    def calcular_total_meses(self, mes_inicial, anio_inicial, mes_final, anio_final):
        return (anio_final - anio_inicial) * 12 + (mes_final - mes_inicial) + 1
    
    def esperar_descarga_completa(self, filepath, timeout=20):
        inicio = time.time()
        while True:
            try:
                with open(filepath, 'rb'):
                    return True
            except (PermissionError, FileNotFoundError):
                if time.time() - inicio > timeout:
                    self.console.log("Tiempo de espera agotado para descarga", "warning")
                    return False
                time.sleep(1)
    
    def abrir_pdf(self):
        if os.path.exists(self.pdf_final_path):
            try:
                if platform.system() == "Windows":
                    os.startfile(self.pdf_final_path)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.run(["open", self.pdf_final_path])
                else:  # Linux
                    subprocess.run(["xdg-open", self.pdf_final_path])
                self.console.log(f"Abriendo PDF unificado: {os.path.basename(self.pdf_final_path)}", "success")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el PDF:\n{e}")
                self.console.log(f"Error abriendo PDF: {e}", "error")
        else:
            messagebox.showwarning("Archivo no encontrado", "No se encontró el PDF unificado")
            self.console.log("PDF unificado no encontrado", "warning")
            
    def run(self):
        self.root.mainloop()
        self.animation_running = False

if __name__ == "__main__":
    app = ANSESDownloaderPro()
    app.run()
